"""ads2coa.py

AIRA INSTITUTE NSF COA Generator

This script generates the NSF COA from the author affiliations csv file.

Steps:
 - Get the author affiliations csv file from ADS and save
   it locally as `authorAffiliations.csv`.

 - Get the unadulterated NSF COA template from the NSF website:
https://www.nsf.gov/bfa/dias/policy/coa/coa_template.xlsx

 - Run this script:
    python ads2coa.py -o lastname_coa.xlsx

 - Edit the generated file and save it as lastname_coa.xlsx
"""

from copy import copy
import pandas as pd
from pathlib import Path
from argparse import ArgumentParser
import os
import urllib.request
import sys

from openpyxl import load_workbook
from openpyxl.utils import rows_from_range
from openpyxl.worksheet.table import Table


TABLE4_START_ROW = 52
TABLE4_TEMPLATE_PERSON = "Alphaman, Lin"
DEFAULT_TEMPLATE_FILENAME = "coa_template.xlsx"
DEFAULT_AUTHOR_AFFILIATIONS_FILENAME = "authorAffiliations.csv"
NUMBER_OF_EXISTING_ROWS_IN_TABLE4 = 5
NSF_TEMPLATE_URL = "https://www.nsf.gov/bfa/dias/policy/coa/coa_template.xlsx"


def copy_range(range_str, sheet, offset):
    """Copy cell values and style to the new row using offset
    https://stackoverflow.com/questions/75390577/python-copy-entire-row-with-formatting
    """
    for row in rows_from_range(range_str):
        for cell in row:
            # if sheet[cell].value is not None:  # Don't copy other cells in merged unit
            dst_cell = sheet[cell].offset(row=offset, column=0)
            src_cell = sheet[cell]

            ### Copy Cell value
            dst_cell.value = src_cell.value

            ### Copy Cell Styles
            dst_cell.font = copy(src_cell.font)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)

            dst_cell.number_format = src_cell.number_format


class COA:
    def __init__(
        self,
        out_filename,
        author_affiliations_filename=DEFAULT_AUTHOR_AFFILIATIONS_FILENAME,
        template_filename=DEFAULT_TEMPLATE_FILENAME,
        dirname=Path.cwd(),
    ):
        """
        Generate the NSF COA from the author affiliations csv file

        Parameters
        ----------
        out_filename : str
            Output filename
        author_affiliations_filename : str
            Author affiliations filename
        template_filname : str
            Template filename
        dirname : str
            Directory name
        """

        if not os.path.exists(template_filename):
            if template_filename == DEFAULT_TEMPLATE_FILENAME:
                print(f"🤔 Could not find `{template_filename}`; downloading default NSF template")
                response = urllib.request.urlretrieve(NSF_TEMPLATE_URL, DEFAULT_TEMPLATE_FILENAME)
            else:
                print(f"🤔 Could not find template `{template_filename}`")
                sys.exit(-1)

        self.wb = load_workbook(template_filename)
        self.wb_alt = load_workbook(template_filename)

        self._check_is_template()

        self.df = pd.read_csv(
            author_affiliations_filename,
            header=None,
            names=["author", "affil", "date"],
            index_col=False,
        )
        self.add_author_affiliations()
        self.wb.save(out_filename)
        print(
            f"🎉 Wrote \033[1m{out_filename}\033[0m. Now edit this and save it as \033[1mlastname_coa.xlsx\033[0m. 🎉"
        )

    def _check_is_template(self):
        """
        Check if the workbook is unaltered template
        """
        if self.wb.sheetnames[0] != "NSF COA Template":
            raise ValueError(
                "The first sheet of the workbook is not the NSF COA template"
            )

        sheet = self.wb["NSF COA Template"]
        if sheet[f"B{TABLE4_START_ROW}"].value != TABLE4_TEMPLATE_PERSON:
            raise ValueError("The first row of the table 4 is not the template person")

        if sheet.tables["TableD"].ref != "A51:E56":
            raise ValueError("The table 4 is not in the correct location")

    def _permute_date(self, date_str):
        """
        Permute the date string to the format of the template

        2022/06/01 -> 06/01/2022

        """
        date = date_str.split("/")
        return f"{date[1]}/{date[2]}/{date[0]}"

    def add_author_affiliations(self):
        """
        Add author affiliations to the template
        """
        sheet = self.wb["NSF COA Template"]
        num_to_add = len(self.df) - NUMBER_OF_EXISTING_ROWS_IN_TABLE4

        # empty the first two rows of table 4
        for i in range(2):  # 2 rows
            for j in range(5):  # 5 columns
                sheet.cell(row=TABLE4_START_ROW + i, column=j + 1).value = None

        # copy the last template row to the end of the new table
        copy_range(
            f"A{TABLE4_START_ROW+5}:E{TABLE4_START_ROW+5}",
            sheet,
            num_to_add + NUMBER_OF_EXISTING_ROWS_IN_TABLE4 + 1,
        )

        # add enough rows to table 4 to accomodate all authors
        sheet.insert_rows(TABLE4_START_ROW + 2, num_to_add)

        for i in range(num_to_add):
            copy_range(f"A{TABLE4_START_ROW}:E{TABLE4_START_ROW}", sheet, 1 + i)

        # add in the authors and their affiliations
        for i, row in self.df.iterrows():
            sheet[f"A{TABLE4_START_ROW+i}"] = "A:"
            sheet[f"B{TABLE4_START_ROW+i}"] = row["author"]
            sheet[f"C{TABLE4_START_ROW+i}"] = row["affil"]
            sheet[f"D{TABLE4_START_ROW+i}"] = ""
            sheet[f"E{TABLE4_START_ROW+i}"] = self._permute_date(row["date"])

        # add the removed tables back in
        del sheet.tables["TableD5"]
        new_d5 = self.wb_alt["NSF COA Template"].tables["TableD5"]
        tab = Table(
            displayName=new_d5.displayName,
            name=new_d5.name,
            ref=f"A{58+num_to_add+NUMBER_OF_EXISTING_ROWS_IN_TABLE4}:E{68+num_to_add+NUMBER_OF_EXISTING_ROWS_IN_TABLE4}",
        )
        sheet.add_table(tab)

        del sheet.tables["TableD"]
        new_d = self.wb_alt["NSF COA Template"].tables["TableD"]
        tab = Table(
            displayName=new_d.displayName,
            name=new_d.name,
            ref=f"A{TABLE4_START_ROW - 1}:E{TABLE4_START_ROW + num_to_add + NUMBER_OF_EXISTING_ROWS_IN_TABLE4 - 1}",
        )
        sheet.add_table(tab)


def main():
    parser = ArgumentParser(
        description="Generate NSF COA from author affiliations csv file"
    )

    parser.add_argument(
        "-o",
        "--out_filename",
        default="coa.xlsx",
        help="Output filename (default: %(default)s)",
    )
    parser.add_argument(
        "-a",
        "--author_affiliations_filename",
        default=DEFAULT_AUTHOR_AFFILIATIONS_FILENAME,
        help="Author affiliations filename (default: %(default)s)",
    )
    parser.add_argument(
        "-t",
        "--template_filename",
        default=DEFAULT_TEMPLATE_FILENAME,
        help="Template filename (default: %(default)s)",
    )
    parser.add_argument(
        "-d",
        "--dirname",
        default=Path.cwd(),
        help="Directory name (default: %(default)s)",
    )

    COA(**vars(parser.parse_args()))


if __name__ == "__main__":
    main()
